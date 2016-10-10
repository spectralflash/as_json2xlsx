# -*- coding: utf-8 -*-
import os
import sys
from numbers import Number
import demjson  # pip install demjson
import codecs
import datetime
import openpyxl
import xlsxwriter
import json
from argparse import ArgumentParser

__author__ = 'Andrey Berestyansky'


def get_params():

    parser = ArgumentParser("Converts pseudo-JSON .as file into XLSX and vice versa.")
    parser.add_argument("file", help=".xlsx or .as file to convert")
    return parser.parse_args()


def save_json_to_xlsx(content, header, xlsx_filename):

    wb = xlsxwriter.Workbook(xlsx_filename)
    ws = wb.add_worksheet()
    ws.title = "Data"
    ws.write(0, 0, header)

    langs = list(content.values()[0].keys())
    for lang in langs:
         ws.write(0, langs.index(lang) + 1, lang)

    str_num = 1
    for string_id in sorted(content.keys()):
        entry = content.pop(string_id)
        ws.write(str_num, 0, string_id)
        for lang in langs:
            entry[lang] = entry[lang].replace('\n', '\\n')
            entry[lang] = entry[lang].replace('\r', '\\r')
            entry[lang] = entry[lang].replace('\t', '\\t')
            entry[lang] = entry[lang].replace('\"', '\\\"')
            ws.write(str_num, langs.index(lang) + 1, entry[lang])
        str_num += 1

    wb.close()


def save_xlsx_to_as(xlsx_filename, as_filename):

    wb = openpyxl.load_workbook(filename=xlsx_filename, data_only=True, read_only=True)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    langs = []
    for row in ws.get_squared_range(2, 1, ws.max_column, 1):
        for cell in row:
            langs.append(cell.value)

    res_dict = {}
    for row in ws.get_squared_range(1, 2, ws.max_column, ws.max_row):
        str_id = str(row[0].value)
        res_dict[str_id] = {}
        lang_idx = 0
        for cell in row[1:]:
            if cell.value is None:
                res_dict[str_id][langs[lang_idx]] = u""
            elif isinstance(cell.value, Number):
                res_dict[str_id][langs[lang_idx]] = str(cell.value).decode('ascii')
            else:
                res_dict[str_id][langs[lang_idx]] = cell.value
            lang_idx += 1

    str_data = ws.cell(None, 1, 1).value + " = {\n"
    for key in sorted(res_dict):
        str_data += "\t" + key + ": {\n"
        for lang in langs:
            str_data += "\t\t" + lang + ": \"" + res_dict[key][lang] + "\",\n"
        str_data = str_data[:-2] + "\n"  # remove last comma
        str_data += "\t},\n"
    str_data = str_data[:-2] + "\n"  # remove last comma
    str_data += "};"

    with codecs.open(as_filename, 'w', 'utf-8') as out_file:
        out_file.write(str_data)


def timestamp():
    return datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


if __name__ == '__main__':

    args = get_params()

    # os.chdir(os.path.dirname(os.path.abspath(args.file)))
    # print os.path.abspath(os.curdir)

    if args.file.endswith(".as"):
        print "Parsing", args.file
        with codecs.open(args.file, 'r', 'utf-8') as as_file:
            as_file_content = as_file.read()
            header_part = as_file_content.split(" = ")[0]
            json_part = as_file_content[len(header_part) + 3:-1]  # -1 to omit the ending ;
            json = demjson.decode(json_part.encode('utf-8'))
        xlsx_filename = args.file.replace('.as', '_' + timestamp() + '.xlsx')
        save_json_to_xlsx(json, header_part, xlsx_filename)
        print ".XLSX saved: " + os.path.abspath(xlsx_filename)

    elif args.file.endswith(".xlsx"):
        print "Parsing", args.file
        as_filename = args.file.replace('.xlsx', '.as')
        save_xlsx_to_as(args.file, as_filename)
        print ".AS saved: " + os.path.abspath(as_filename)

    else:
        sys.exit("Unknown file format. Requires an XLSX or AS file.")
